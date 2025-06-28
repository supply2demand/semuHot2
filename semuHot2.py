import sys
import os
import json
import keyboard  # 핫스트링 처리 (관리자 권한 필요할 수 있음)
from openpyxl import load_workbook

# PyQt6 임포트
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, QLabel, QHeaderView,
    QToolButton, QDialog, QFormLayout, QKeySequenceEdit, QInputDialog, QMenu,
    QCheckBox, QStyle, QSizePolicy
)
from PyQt6.QtGui import QShortcut, QKeySequence, QDesktopServices, QIcon, QPainter, QPixmap
from PyQt6.QtCore import Qt, QUrl, QRect

# ---------------------------
# PyInstaller로 exe를 만들 때 이미지가 보이지 않는 경우 참고
# (1) --add-data "이미지.png;." 등으로 빌드 시 옵션을 정확히 지정해야 합니다.
# (2) code 내에서 frozen 상태일 때 sys._MEIPASS를 통해 실제 이미지 경로를 추적해야 합니다.
# ---------------------------

# ---------------------------
# 실행 파일과 같은 폴더 경로 설정
# ---------------------------
if getattr(sys, 'frozen', False):
    script_dir = sys._MEIPASS
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

current_json_file = os.path.join(script_dir, "data.json")   # (데이터 저장용)
shortcuts_file = os.path.join(script_dir, "shortcuts.json")

# ---------------------------
# 전역 변수
# ---------------------------
current_mode = "법인"     # 기본 토글 모드("법인" / "개인")
hotstring_active = False  # 핫스트링 활성 여부

# 핫스트링 안내 메시지 표시 여부 추적
have_shown_enable_message = False
have_shown_disable_message = False

# 엑셀(또는 JSON)에서 읽어온 데이터: 두 시트를 분리해서 저장
corp_data = []      # [{ "지정": str, "번호": int, "구분": str }, ... ] - 법인 시트
personal_data = []  # [{ "지정": str, "번호": int, "구분": str }, ... ] - 개인 시트

# 현재 등록된 abbreviation을 추적 (최적화용)
active_abbreviations = set()

# ---------------------------
# QKeyCombination 대체용 폴백 함수 (구버전 PyQt6, PyQt5 호환)
# ---------------------------
def make_keysequence_from_event(event):
    modifiers = event.modifiers()
    key = event.key()
    chord = 0
    if modifiers & Qt.KeyboardModifier.ControlModifier:
        chord |= Qt.KeyboardModifier.ControlModifier.value
    if modifiers & Qt.KeyboardModifier.ShiftModifier:
        chord |= Qt.KeyboardModifier.ShiftModifier.value
    if modifiers & Qt.KeyboardModifier.AltModifier:
        chord |= Qt.KeyboardModifier.AltModifier.value
    if modifiers & Qt.KeyboardModifier.MetaModifier:
        chord |= Qt.KeyboardModifier.MetaModifier.value
    chord |= key
    return QKeySequence(chord)

# ---------------------------
# TransparentTableWidget 클래스
# ---------------------------
class TransparentTableWidget(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Hyung.png 파일을 배경 이미지로 로드합니다.
        bg_image_path = os.path.join(script_dir, "Hyung.png")
        self.bg_pixmap = QPixmap(bg_image_path)
        # 테이블의 배경을 투명하게 만들기 위해 기본 배경은 제거합니다.
        self.setStyleSheet("background: transparent;")
    
    def paintEvent(self, event):
        painter = QPainter(self.viewport())
        # 10% 투명도로 배경 이미지 그리기
        painter.setOpacity(0.1)
        rect = self.viewport().rect()
        # 원본 이미지 비율을 유지하면서 viewport 크기에 맞게 스케일
        scaled_pixmap = self.bg_pixmap.scaled(rect.size(), Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        # 중앙에 배치
        x = (rect.width() - scaled_pixmap.width()) // 2
        y = (rect.height() - scaled_pixmap.height()) // 2
        painter.drawPixmap(x, y, scaled_pixmap)
        painter.setOpacity(1.0)
        # 기본 그리기 (셀 및 내용)
        super().paintEvent(event)

# ---------------------------
# 토글 스위치 (법인/개인)
# ---------------------------
class ToggleSwitch(QCheckBox):
    def __init__(self, on_text="ON", off_text="OFF", parent=None):
        super().__init__(off_text, parent)
        self.on_text = on_text
        self.off_text = off_text
        self.setChecked(False)
        self.setText(self.off_text)
        self.setStyleSheet("""
            QCheckBox::indicator {
                width: 40px; height: 20px;
            }
            QCheckBox::indicator:checked {
                background-color: #8e44ad;
                border-radius: 10px;
            }
            QCheckBox::indicator:unchecked {
                background-color: #cccccc;
                border-radius: 10px;
            }
        """)
        self.stateChanged.connect(self.on_state_changed)

    def on_state_changed(self, state):
        if self.isChecked():
            self.setText(self.on_text)
        else:
            self.setText(self.off_text)

# ---------------------------
# 클릭 시 URL 열리는 라벨
# ---------------------------
class ClickableLabel(QLabel):
    def __init__(self, text, link, parent=None):
        super().__init__(text, parent)
        self.link = link
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.setWordWrap(True)
        self.setStyleSheet("margin: 5px; color: #2c3e50;")

    def mousePressEvent(self, event):
        QDesktopServices.openUrl(QUrl(self.link))
        event.accept()

# ---------------------------
# 한 개 키(chord)만 입력받는 KeySequenceEdit
# ---------------------------
class SingleChordKeySequenceEdit(QKeySequenceEdit):
    def keyPressEvent(self, event):
        if event.key() == Qt.Key.Key_Backspace:
            self.clear()
            event.accept()
            return
        if event.key() in (Qt.Key.Key_Control, Qt.Key.Key_Shift, Qt.Key.Key_Alt, Qt.Key.Key_Meta):
            event.accept()
            return
        try:
            from PyQt6.QtGui import QKeyCombination
            keyseq = QKeySequence(QKeyCombination(event.modifiers(), event.key()))
        except ImportError:
            keyseq = make_keysequence_from_event(event)
        self.setKeySequence(keyseq)
        event.accept()

# ---------------------------
# 단축키 설정 다이얼로그
# ---------------------------
class SettingsDialog(QDialog):
    def __init__(self, current_shortcuts, parent=None):
        super().__init__(parent)
        self.setWindowTitle("단축키 설정")
        self.current_shortcuts = current_shortcuts
        layout = QFormLayout(self)
        self.edits = {}
        for key, seq in current_shortcuts.items():
            key_edit = SingleChordKeySequenceEdit(seq)
            layout.addRow(f"{key}:", key_edit)
            self.edits[key] = key_edit

        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("확인")
        btn_cancel = QPushButton("취소")
        btn_ok.clicked.connect(self.accept)
        btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(btn_ok)
        btn_layout.addWidget(btn_cancel)
        layout.addRow(btn_layout)

    def get_shortcuts(self):
        new_shortcuts = {}
        for key, widget in self.edits.items():
            new_shortcuts[key] = widget.keySequence()
        return new_shortcuts

# ---------------------------
# 메인 윈도우
# ---------------------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("세무사랑 핫스트링")
        self.resize(800, 600)

        # 프로그램 아이콘 설정 (같은 폴더의 Hyung.png)
        program_icon_path = os.path.join(script_dir, "Hyung.png")
        self.setWindowIcon(QIcon(program_icon_path))

        # 기본 단축키
        default_shortcuts = {
            "불러오기": "",
            "핫스트링 활성화": "",
            "현재 카테고리": "",
            "찾기": "Ctrl+F",
            "프로그램 종료": "Ctrl+Q"
        }
        if os.path.exists(shortcuts_file):
            try:
                with open(shortcuts_file, "r", encoding="utf-8") as f:
                    loaded = json.load(f)
                self.shortcuts = {}
                for key in default_shortcuts:
                    seq_str = loaded.get(key, default_shortcuts[key])
                    self.shortcuts[key] = QKeySequence(seq_str) if seq_str else QKeySequence()
            except Exception:
                self.shortcuts = {k: QKeySequence(v) if v else QKeySequence() for k, v in default_shortcuts.items()}
        else:
            self.shortcuts = {k: QKeySequence(v) if v else QKeySequence() for k, v in default_shortcuts.items()}

        self.updating_table = False

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        # 상단 버튼/스위치 레이아웃
        top_layout = QHBoxLayout()
        main_layout.addLayout(top_layout)

        # [불러오기] 버튼
        self.btn_load = QPushButton("불러오기")
        load_menu = QMenu()
        action_excel = load_menu.addAction("엑셀 불러오기")
        action_json = load_menu.addAction("JSON 불러오기")
        action_excel.triggered.connect(self.load_excel)
        action_json.triggered.connect(self.load_json_file)
        self.btn_load.setMenu(load_menu)
        top_layout.addWidget(self.btn_load)

        # 핫스트링 스위치
        self.switch_hotstring = ToggleSwitch(on_text="활성화", off_text="비활성화")
        self.switch_hotstring.setChecked(hotstring_active)
        self.switch_hotstring.stateChanged.connect(self.on_hotstring_switch_changed)
        top_layout.addWidget(self.switch_hotstring)

        # 카테고리 스위치 (법인/개인)
        self.switch_category = ToggleSwitch(on_text="법인", off_text="개인")
        self.switch_category.setChecked(current_mode == "법인")
        self.switch_category.stateChanged.connect(self.on_category_switch_changed)
        top_layout.addWidget(self.switch_category)

        # [설정] 아이콘 버튼 - 같은 폴더의 settingIcon.png 사용
        top_layout.addStretch()
        self.btn_settings = QToolButton()
        setting_icon_path = os.path.join(script_dir, "settingIcon.png")
        self.btn_settings.setIcon(QIcon(setting_icon_path))
        self.btn_settings.setToolTip("단축키 설정")
        self.btn_settings.clicked.connect(self.open_settings_dialog)
        top_layout.addWidget(self.btn_settings)

        # 테이블 - TransparentTableWidget 사용 (배경에 투명 이미지)
        self.table = TransparentTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["지정", "번호", "구분"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked | QTableWidget.EditTrigger.SelectedClicked)
        self.table.cellChanged.connect(self.on_cell_changed)
        main_layout.addWidget(self.table)

        # 하단 라벨
        author_label = ClickableLabel("by 박종석 세무회계", "https://blog.naver.com/jstax220")
        author_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        main_layout.addWidget(author_label)

        # QShortcut 설정
        self.shortcut_load = QShortcut(self.shortcuts["불러오기"], self)
        self.shortcut_load.activated.connect(self.show_load_menu)
        self.shortcut_hotstring = QShortcut(self.shortcuts["핫스트링 활성화"], self)
        self.shortcut_hotstring.activated.connect(self.toggle_hotstring_via_shortcut)
        self.shortcut_category = QShortcut(self.shortcuts["현재 카테고리"], self)
        self.shortcut_category.activated.connect(self.toggle_category_via_shortcut)
        self.shortcut_find = QShortcut(self.shortcuts["찾기"], self)
        self.shortcut_find.activated.connect(self.search_table)
        self.shortcut_exit = QShortcut(self.shortcuts["프로그램 종료"], self)
        self.shortcut_exit.activated.connect(self.exit_program)

        if os.path.exists(current_json_file):
            self.load_json_data()

    def show_load_menu(self):
        pos = self.btn_load.mapToGlobal(self.btn_load.rect().bottomLeft())
        self.btn_load.menu().exec(pos)

    def on_hotstring_switch_changed(self, _state):
        global hotstring_active
        if self.switch_hotstring.isChecked():
            self.enable_hotstring()
        else:
            self.disable_hotstring()

    def on_category_switch_changed(self, _state):
        global current_mode
        current_mode = "법인" if self.switch_category.isChecked() else "개인"
        self.update_table()
        if hotstring_active:
            self.update_hotstrings()

    def enable_hotstring(self):
        global hotstring_active, have_shown_enable_message
        if not hotstring_active:
            hotstring_active = True
            self.update_hotstrings()
            if not have_shown_enable_message:
                QMessageBox.information(self, "정보", "핫스트링 기능이 활성화되었습니다.\n(입력 후 스페이스바 누르면 치환됩니다.)")
                have_shown_enable_message = True

    def disable_hotstring(self):
        global hotstring_active, active_abbreviations, have_shown_disable_message
        if hotstring_active:
            for ab in active_abbreviations:
                try:
                    keyboard.remove_abbreviation(ab)
                except:
                    pass
            active_abbreviations.clear()
            hotstring_active = False
            if not have_shown_disable_message:
                QMessageBox.information(self, "정보", "핫스트링 기능이 비활성화되었습니다.")
                have_shown_disable_message = True

    def toggle_hotstring_via_shortcut(self):
        self.switch_hotstring.setChecked(not self.switch_hotstring.isChecked())

    def toggle_category_via_shortcut(self):
        self.switch_category.setChecked(not self.switch_category.isChecked())

    def update_hotstrings(self):
        global hotstring_active, active_abbreviations, corp_data, personal_data, current_mode
        if not hotstring_active:
            return
        target_data = corp_data if current_mode == "법인" else personal_data
        new_abbrevs = set(item["지정"] for item in target_data)
        to_remove = active_abbreviations - new_abbrevs
        to_add = new_abbrevs - active_abbreviations
        for ab in to_remove:
            try:
                keyboard.remove_abbreviation(ab)
            except:
                pass
        for ab in to_add:
            for row in target_data:
                if row["지정"] == ab:
                    keyboard.add_abbreviation(ab, str(row["번호"]))
                    break
        active_abbreviations = (active_abbreviations - to_remove) | to_add

    def load_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "엑셀 파일 선택",
            script_dir,
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if not file_path:
            return
        try:
            wb = load_workbook(file_path, data_only=True)
            temp_corp = []
            temp_personal = []
            if "법인" not in wb.sheetnames:
                QMessageBox.warning(self, "경고", "엑셀 파일에 '법인' 시트가 없습니다.")
            else:
                ws_corp = wb["법인"]
                for row in ws_corp.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    지정 = str(row[0]).strip()
                    번호 = row[1]
                    구분 = str(row[2]).strip() if row[2] else ""
                    try:
                        번호 = int(번호)
                    except:
                        continue
                    temp_corp.append({"지정": 지정, "번호": 번호, "구분": 구분})
            if "개인" not in wb.sheetnames:
                QMessageBox.warning(self, "경고", "엑셀 파일에 '개인' 시트가 없습니다.")
            else:
                ws_personal = wb["개인"]
                for row in ws_personal.iter_rows(min_row=2, values_only=True):
                    if row[0] is None:
                        break
                    지정 = str(row[0]).strip()
                    번호 = row[1]
                    구분 = str(row[2]).strip() if row[2] else ""
                    try:
                        번호 = int(번호)
                    except:
                        continue
                    temp_personal.append({"지정": 지정, "번호": 번호, "구분": 구분})
            data_to_save = {"corp_data": temp_corp, "personal_data": temp_personal}
            with open(current_json_file, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=4)
            self.load_json_data()
        except Exception as e:
            QMessageBox.critical(self, "오류", str(e))

    def load_json_file(self):
        global corp_data, personal_data, current_json_file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "JSON 파일 선택",
            script_dir,
            "JSON Files (*.json);;All Files (*)"
        )
        if not file_path:
            return
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            corp_data = data.get("corp_data", [])
            personal_data = data.get("personal_data", [])
            current_json_file = file_path
            self.update_table()
            if hotstring_active:
                self.update_hotstrings()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"JSON 로드 실패: {str(e)}")

    def load_json_data(self):
        global corp_data, personal_data
        if not os.path.exists(current_json_file):
            QMessageBox.warning(self, "경고", f"JSON 파일({current_json_file})이 존재하지 않습니다.")
            return
        try:
            with open(current_json_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            corp_data = data.get("corp_data", [])
            personal_data = data.get("personal_data", [])
            self.update_table()
            if hotstring_active:
                self.update_hotstrings()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"JSON 로드 실패: {str(e)}")

    def update_table(self):
        global current_mode, corp_data, personal_data
        self.updating_table = True
        self.table.blockSignals(True)
        show_data = corp_data if current_mode == "법인" else personal_data
        self.table.setRowCount(len(show_data))
        for row_idx, entry in enumerate(show_data):
            item_지정 = QTableWidgetItem(str(entry["지정"]))
            item_번호 = QTableWidgetItem(str(entry["번호"]))
            item_구분 = QTableWidgetItem(str(entry["구분"]))
            item_지정.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_번호.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item_구분.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 0, item_지정)
            self.table.setItem(row_idx, 1, item_번호)
            self.table.setItem(row_idx, 2, item_구분)
        self.table.blockSignals(False)
        self.updating_table = False

    def on_cell_changed(self, row, column):
        if self.updating_table:
            return
        global current_mode, corp_data, personal_data, hotstring_active
        target_data = corp_data if current_mode == "법인" else personal_data
        if row < 0 or row >= len(target_data):
            return
        new_value = self.table.item(row, column).text()
        if column == 0:
            target_data[row]["지정"] = new_value.strip()
        elif column == 1:
            try:
                new_num = int(new_value)
            except:
                QMessageBox.critical(self, "오류", "번호는 정수여야 합니다.")
                self.update_table()
                return
            target_data[row]["번호"] = new_num
        elif column == 2:
            target_data[row]["구분"] = new_value.strip()
        if hotstring_active:
            self.update_hotstrings()

    def search_table(self):
        search_text, ok = QInputDialog.getText(self, "찾기", "찾을 값 입력:")
        if ok and search_text:
            items = self.table.findItems(search_text, Qt.MatchFlag.MatchContains)
            if items:
                item = items[0]
                self.table.setCurrentItem(item)
                self.table.scrollToItem(item)
            else:
                QMessageBox.information(self, "찾기", f"'{search_text}'(을)를 찾을 수 없습니다.")

    def exit_program(self):
        self.close()

    def open_settings_dialog(self):
        dialog = SettingsDialog(self.shortcuts, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_shortcuts = dialog.get_shortcuts()
            self.shortcuts.update(new_shortcuts)
            self.shortcut_load.setKey(self.shortcuts["불러오기"])
            self.shortcut_hotstring.setKey(self.shortcuts["핫스트링 활성화"])
            self.shortcut_category.setKey(self.shortcuts["현재 카테고리"])
            self.shortcut_find.setKey(self.shortcuts["찾기"])
            self.shortcut_exit.setKey(self.shortcuts["프로그램 종료"])
            self.save_shortcut_settings()

    def save_shortcut_settings(self):
        save_dict = {key: self.shortcuts[key].toString() for key in self.shortcuts}
        try:
            with open(shortcuts_file, "w", encoding="utf-8") as f:
                json.dump(save_dict, f, ensure_ascii=False, indent=4)
        except Exception as e:
            QMessageBox.critical(self, "오류", f"단축키 설정 저장 실패: {str(e)}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet("""
    QWidget {
        background-color: #f7f7f7;
        font-family: "Malgun Gothic", "돋움", sans-serif;
        font-size: 11pt;
        color: #2c3e50;
    }
    QPushButton {
        background-color: #8e44ad;
        color: #ecf0f1;
        border: none;
        border-radius: 4px;
        padding: 8px 16px;
    }
    QPushButton:hover {
        background-color: #9b59b6;
    }
    QPushButton:pressed {
        background-color: #7d3c98;
    }
    QCheckBox {
        color: #8e44ad;
        font-weight: bold;
    }
    QCheckBox::indicator {
        width: 40px;
        height: 20px;
    }
    QToolButton {
        background: transparent;
        border: none;
    }
    QTableWidget {
        background-color: #ffffff;
        gridline-color: #bdc3c7;
        color: #2c3e50;
    }
    QHeaderView::section {
        background-color: #8e44ad;
        color: #ecf0f1;
        padding: 6px;
        border: none;
    }
    QLabel {
        color: #8e44ad;
    }
    QMenu {
        background-color: #ffffff;
        color: #2c3e50;
        border: 1px solid #bdc3c7;
    }
    QMenu::item:selected {
        background-color: #8e44ad;
        color: #ecf0f1;
    }
    """)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
